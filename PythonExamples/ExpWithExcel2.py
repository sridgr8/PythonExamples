'''
Created on May 5, 2019

@author: Srinivasulu
'''
import openpyxl    

filename = 'hello.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb.active
rowNum=ws.max_row+1
ws.cell(column=1, row=rowNum, value="Hello")
ws.cell(column=2, row=rowNum, value="World")
wb.save(filename)
