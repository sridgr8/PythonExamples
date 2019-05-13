'''
Created on May 5, 2019

@author: Srinivasulu
'''
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep
import openpyxl    


print("Execution Started")
driver=webdriver.Chrome(executable_path="D:\\SELENIUM\\Chrome Driver\\chromedriver.exe")
driver.implicitly_wait(10)
driver.maximize_window()
driver.get("http://192.168.0.104:8080/job/Hello%20World/20/console")

# filename = 'D:\ECLIPSE PROJECTS\TestRepoOne\PythonExamples\hello.xlsx'
# wb = openpyxl.load_workbook(filename)
# ws = wb.active
# rowNum=ws.max_row+1
# ws.cell(column=1, row=rowNum, value="Hello")
# ws.cell(column=2, row=rowNum, value="World")
# wb.save(filename)
print("Execution Completed")
