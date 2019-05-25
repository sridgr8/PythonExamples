'''
Created on May 5, 2019

@author: Srinivasulu
'''

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from time import sleep
import openpyxl 
import sys
import string

testPlanName = sys.argv[1]
buildUrl = sys.argv[2]
executiondate=sys.argv[3]
buildUser=sys.argv[4]


print("Execution Started")
driver=webdriver.Chrome(executable_path="D:\\SELENIUM\\Chrome Driver\\chromedriver.exe")
driver.implicitly_wait(10)
driver.maximize_window()
driver.get(buildUrl+"timestamps/?elapsed=HH:mm:ss.S")

txtUN=driver.find_element_by_name("j_username")
txtUN.clear()
txtUN.send_keys("sridgr8")

txtPW=driver.find_element_by_name("j_password")
txtPW.clear()
txtPW.send_keys("purgatory")

txtPW.send_keys(Keys.RETURN)

txtTimeContent=str(driver.find_element_by_tag_name("pre").text)
txtTimeElapsed=txtTimeContent.splitlines()[-1]


print(txtTimeContent)
print("Total Time: "+txtTimeElapsed)

driver.get(buildUrl+"logText/progressiveText?start=0")

txtContent=str(driver.find_element_by_tag_name("pre").text)
txtContentLastLine=txtContent.splitlines()[-1]
txtBuildStatus=txtContentLastLine.split()[-1]
print("Build Status: "+txtBuildStatus)

filename = 'D:\\ECLIPSE PROJECTS\\TestRepoOne\\PythonExamples\\hello.xlsx'
wb = openpyxl.load_workbook(filename)
ws = wb.active
rowNum=ws.max_row+1
 
ws.cell(column=1, row=rowNum, value=testPlanName)
ws.cell(column=2, row=rowNum, value=txtTimeElapsed)
ws.cell(column=3, row=rowNum, value=executiondate)
ws.cell(column=4, row=rowNum, value=txtBuildStatus)
ws.cell(column=5, row=rowNum, value=buildUser)
wb.save(filename)

print("Execution Completed")
